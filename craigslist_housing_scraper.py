#! python3
# craigslist_housing_scraper.py -- Searches Craigslist housing for a city and stores links, date & time of posting, titles, and prices to each posting in an xls workbook.

import requests, bs4, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#Asks user for city of Craigslist housing search.
print('Input city: ')
city = input()

#Uses user input to set the starting url.
url = 'https://'+ city + '.craigslist.org/d/housing/search/hhh'

#Creates workbook. 
wb = Workbook() 

#Gets sheet. 
sheet = wb.active

#Get page from link and convert to BeautifulSoup for html parsing.
def get_soup(url):
	res = requests.get(url)
	res.raise_for_status()
	soup = bs4.BeautifulSoup(res.text, features="lxml")
	return soup

#Returns a new row from the workbook sheet.
def get_new_row(sheet):	
	new_row = sheet.max_row
	return new_row

#Writes info to workbook.
def write_to_excel(row_index, column_index, text):
	sheet.cell(row=row_index, column=column_index).value = text	

#Scrapes listings for post links.		
def scrape_links(element, soup, sheet, new_row, column):
	elem = soup.select(element)
	num = len(elem)		
	for i in range(num):
		text = elem[i].get('href')
		new_row = get_new_row(sheet) + 1
		write_to_excel(new_row, column, text)

#Goes to next page and gets new url.
def get_new_url(soup):
	new = soup.find(class_="button next")
	url_end = str(new.get('href'))
	url = 'https://' + city + '.craigslist.org/d/housing' + url_end
	print('Scraping post links for: ' + url)	
	return url

#Scrapes listings.
def scrape(element, soup, sheet, url_row, column):
	elem = soup.select(element)
	num = len(elem)		
	for i in range(num):
		text = elem[i].get_text(element)
		#print(text)
		row = url_row
		write_to_excel(row, column, text)

new_row = 0 #Sets row at 0 for the purpose of the loop.

#65536 is the maximum number of rows Excel can hold in a single sheet. While the maximum number of rows is less than this limit, program will loop, scraping each page for links from the Craigslist listings.
while new_row < 65536:

	try:
		
		#Calls the function 'get_soup' and sets this 'soup' variable to its return, which is the page from the url converted to BeautifulSoup for html parsing.
		soup = get_soup(url)		
		#Scrapes for links using the inspected element class of the link.	
		scrape_links('a.result-title.hdrlnk', soup, sheet, new_row, 1)
		#Calls the function 'get_new_url' and sets this 'url' variable to its return, which is the url of the next page of Craigslist listings.
		url = get_new_url(soup)

	#Ends the loop when a url is invalid.	
	except requests.exceptions.RequestException:
		#Displays completion of scraping post links and states how many posts were found.
		print('Completed scraping post links. ' + str(sheet.max_row) + ' posts found.')
		break

i = 1 #sets start of i for the loop.

#Loops through each row in the first column of the workbook, which is where our links to each Craigslist posting are stored.
while i < sheet.max_row + 1:
	try:

		#Sets url to the next Craigslist posting link, which is stored in the first column of our workbook.
		url = sheet.cell(row=i, column=1).value

		#Calls the function 'get_soup' and sets this 'soup' variable to its return, which is the page from the url converted to BeautifulSoup for html parsing.
		soup = get_soup(url)

		#Calculates percentage of post scraping completion.
		percentage_complete = i/sheet.max_row * 100
		#Displays percentage of post scraping completion.
		print('Scraping posts. ' + str(percentage_complete) + '% complete.')
		
		#Scrapes each posting for date and time posted using inspected element id and tag.		
		scrape('#display-date time', soup, sheet, i, 2)
		#Scrapes each posting for titles using inspected element id.		
		scrape('#titletextonly', soup, sheet, i, 3)
		#Scrapes each posting for prices using inspected element class.
		scrape('.price', soup, sheet, i, 4)
		
		i = i + 1 #Adds one to i in order to loop to the next row of urls.

	#Ends the program when a url is invalid.	
	except requests.exceptions.RequestException:
		print('Done.')
		break	

#Saves the final workbook.
wb.save('craigslist.xls')
print('Workbook saved.')
    